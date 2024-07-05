
from openpyxl import load_workbook

class Excel_Functions:

    def __init__(self, excel_file_name, sheet_name):
        self.file = excel_file_name
        self.sheet = sheet_name

    # Get the Row Count of the Excel Sheet
    def Row_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_row)

    # Get the column count of the Excel Sheet
    def Column_Count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.max_column)

    # Read the data from the Excel File
    def Read_Data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return (sheet.cell(row=row_number, column=column_number).value)

    # Write the data into the Excel File
    def Write_Data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)
